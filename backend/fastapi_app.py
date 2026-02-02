#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from fastapi import FastAPI, File, UploadFile, HTTPException, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
from pydantic import BaseModel, Field
from typing import Optional, List, Dict, Any
import os
import logging
import json
import csv
from datetime import datetime, date
from rag_rga import RAGService, DocumentProcessor
from weather_openmeteo import obtener_clima, validar_fecha
from workflow_trigger import trigger_workflow
from marwis import run_marwis
from avisos import generar_avisos, obtener_tareas_procedimiento
from simulacion import generar_datos_simulados, obtener_escenarios_disponibles
from isuite import enviar_aviso_a_isuite, verificar_configuracion as verificar_config_isuite, obtener_estado_token

# Configuración de logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configuración de la aplicación
app = FastAPI(
    title="Aeropuertos API",
    description="API para análisis climático y gestión de procedimientos aeroportuarios",
    version="1.0.0"
)

# Configuración CORS
allowed_origin = os.getenv("ALLOWED_ORIGIN")
cors_origins = [allowed_origin] if allowed_origin else ["*"]
app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins,  # En producción, especifica los orígenes permitidos
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configuración
UPLOAD_FOLDER = 'uploads'
MAX_FILE_SIZE = 16 * 1024 * 1024  # 16MB
ALLOWED_EXTENSIONS = {'pdf', 'docx', 'xlsx', 'xls', 'csv', 'pptx', 'ppt'}

# Asegurar que existe la carpeta de uploads
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ==================== MODELOS PYDANTIC ====================

class QuestionRequest(BaseModel):
    question: str = Field(..., min_length=1, description="Pregunta para el chatbot")

class WeatherRequest(BaseModel):
    ciudad: str = Field(default="rio grande", description="Ciudad para consultar el clima")
    fecha: str = Field(..., description="Fecha en formato YYYY-MM-DD")

class QuestionResponse(BaseModel):
    success: bool
    answer: Optional[str] = None
    sources: Optional[List[Dict[str, Any]]] = None
    message: Optional[str] = None

class WeatherResponse(BaseModel):
    success: bool
    resultado: Optional[Dict[str, Any]] = None
    message: Optional[str] = None

class StatsResponse(BaseModel):
    success: bool
    stats: Optional[Dict[str, Any]] = None
    message: Optional[str] = None

class UploadResponse(BaseModel):
    success: bool
    message: str

class ClearResponse(BaseModel):
    success: bool
    message: str

class HistoricoRequest(BaseModel):
    fecha_inicio: str = Field(..., description="Fecha de inicio en formato YYYY-MM-DD")
    fecha_fin: str = Field(..., description="Fecha de fin en formato YYYY-MM-DD")
    limite: Optional[int] = Field(default=1000, description="Número máximo de registros a retornar")

class HistoricoResponse(BaseModel):
    success: bool
    datos: Optional[List[Dict[str, Any]]] = None
    total_registros: Optional[int] = None
    message: Optional[str] = None

# ==================== MODELOS MARWIS ====================

class SensorChannel(BaseModel):
    SensorChannelId: Optional[str] = None
    SensorChannelName: Optional[str] = None
    SensorChannelUnit: Optional[str] = None
    SensorTypeId: Optional[int] = None
    SensorIndex: Optional[int] = None
    ValueType: Optional[int] = None

class StationSensorsResponse(BaseModel):
    success: bool
    sensors: Optional[List[SensorChannel]] = None
    total: Optional[int] = None
    message: Optional[str] = None

# ==================== FUNCIONES AUXILIARES ====================

def allowed_file(filename: str) -> bool:
    """Verificar si el archivo tiene una extensión permitida"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def analizar_condiciones_climaticas(datos_clima: Dict[str, Any]) -> Dict[str, Any]:
    """Analiza las condiciones climáticas y extrae información relevante"""
    try:
        location = datos_clima.get('location', {})
        current = datos_clima.get('current', {})
        forecast = datos_clima.get('forecast', {}).get('forecastday', [])
        
        # Obtener temperatura - soportar tanto temp_c directo como dentro de current
        temp_actual = current.get('temp_c')
        if temp_actual is None:
            temp_actual = 0
        
        humedad = current.get('humidity', 0)
        
        # Obtener viento - puede ser None en algunas estaciones PWS
        viento_kmh = current.get('wind_kph')
        if viento_kmh is None:
            viento_kmh = 0
        
        visibilidad = current.get('vis_km', None)
        
        # Obtener precipitación - puede ser None en algunas estaciones PWS
        precipitacion = current.get('precip_mm')
        if precipitacion is None:
            precipitacion = 0
        
        condicion = current.get('condition', {}).get('text', 'Datos de estación meteorológica')
        
        pronostico_dia = {}
        if forecast:
            day_data = forecast[0].get('day', {})
            pronostico_dia = {
                'temp_max': day_data.get('maxtemp_c', 0),
                'temp_min': day_data.get('mintemp_c', 0),
                'prob_lluvia': day_data.get('daily_chance_of_rain', 0),
                'prob_nieve': day_data.get('daily_chance_of_snow', 0),
                'precipitacion_total': day_data.get('totalprecip_mm', 0),
                'viento_max': day_data.get('maxwind_kph', 0)
            }
        
        condiciones_adversas = []
        
        if temp_actual is not None and temp_actual < 0:
            condiciones_adversas.append("temperatura bajo cero")
        elif temp_actual is not None and temp_actual > 30:
            condiciones_adversas.append("temperatura alta")
        
        if viento_kmh and viento_kmh > 50:
            condiciones_adversas.append("viento muy fuerte")
        elif viento_kmh and viento_kmh > 30:
            condiciones_adversas.append("viento fuerte")
        
        if visibilidad is not None and visibilidad < 5:
            condiciones_adversas.append("visibilidad reducida")
        
        if (precipitacion and precipitacion > 0) or pronostico_dia.get('prob_lluvia', 0) > 50:
            condiciones_adversas.append("lluvia")
        
        if pronostico_dia.get('prob_nieve', 0) > 30:
            condiciones_adversas.append("nieve")
        
        # Obtener nombre de ubicación - soportar formato PWS
        ubicacion_nombre = location.get('name', '')
        ubicacion_region = location.get('region', '')
        ubicacion_country = location.get('country', '')
        
        if ubicacion_region:
            ubicacion_str = f"{ubicacion_nombre}, {ubicacion_region}"
        elif ubicacion_country:
            ubicacion_str = f"{ubicacion_nombre}, {ubicacion_country}"
        else:
            ubicacion_str = ubicacion_nombre
        
        # Obtener fecha - soportar formato PWS (last_updated en current)
        fecha_str = ''
        if location.get('localtime'):
            fecha_str = location.get('localtime', '').split(' ')[0]
        elif current.get('last_updated'):
            fecha_str = current.get('last_updated', '').split(' ')[0]
        
        return {
            'ubicacion': ubicacion_str,
            'temperatura_actual': temp_actual,
            'condicion_actual': condicion,
            'viento': viento_kmh if viento_kmh else 'N/A',
            'visibilidad': visibilidad if visibilidad is not None else 'N/A',
            'humedad': humedad,
            'precipitacion': precipitacion if precipitacion else 0,
            'pronostico': pronostico_dia,
            'condiciones_adversas': condiciones_adversas,
            'fecha': fecha_str
        }
        
    except Exception as e:
        logger.error(f"Error analizando condiciones climáticas: {e}")
        return {}

def generar_consulta_rag(condiciones: Dict[str, Any]) -> str:
    """Genera una consulta contextual para el sistema RAG"""
    temp_actual = condiciones.get('temperatura_actual', 0)
    condicion_actual = condiciones.get('condicion_actual', '')
    condiciones_adversas = condiciones.get('condiciones_adversas', [])
    viento = condiciones.get('viento', 0)
    visibilidad = condiciones.get('visibilidad', 0)
    prob_lluvia = condiciones.get('pronostico', {}).get('prob_lluvia', 0)
    prob_nieve = condiciones.get('pronostico', {}).get('prob_nieve', 0)
    
    contexto_climatico = f"""
    ANÁLISIS METEOROLÓGICO ACTUAL:
    - Temperatura: {temp_actual}°C
    - Condición: {condicion_actual}
    - Viento: {viento} km/h
    - Visibilidad: {visibilidad} km
    - Probabilidad de lluvia: {prob_lluvia}%
    - Probabilidad de nieve: {prob_nieve}%
    - Condiciones adversas detectadas: {', '.join(condiciones_adversas) if condiciones_adversas else 'Ninguna'}
    
    CONSULTA ESPECÍFICA:
    """
    
    if condiciones_adversas:
        if "lluvia" in condiciones_adversas:
            consulta_especifica = "Necesito procedimientos específicos para operaciones con lluvia. ¿Qué tareas aeroportuarias se ven afectadas por la lluvia y cuáles son los protocolos a seguir?"
        elif "nieve" in condiciones_adversas or prob_nieve > 30:
            consulta_especifica = "Se detectaron condiciones de nieve o alta probabilidad de nieve. ¿Cuáles son los procedimientos para hielo y nieve en pista, incluyendo aplicación de descongelantes?"
        elif "viento fuerte" in condiciones_adversas or "viento muy fuerte" in condiciones_adversas:
            consulta_especifica = f"Se detectó viento fuerte ({viento} km/h). ¿Cuáles son las limitaciones operativas y procedimientos de seguridad para viento fuerte?"
        elif "visibilidad reducida" in condiciones_adversas:
            consulta_especifica = f"La visibilidad está reducida a {visibilidad} km. ¿Cuáles son los procedimientos para operaciones con baja visibilidad?"
        elif "temperatura bajo cero" in condiciones_adversas:
            consulta_especifica = f"La temperatura está bajo cero ({temp_actual}°C). ¿Cuáles son los procedimientos preventivos para evitar formación de hielo?"
        else:
            consulta_especifica = f"Se detectaron condiciones adversas: {', '.join(condiciones_adversas)}. ¿Cuáles son los procedimientos operativos correspondientes?"
    else:
        consulta_especifica = f"""
        Las condiciones meteorológicas son NORMALES:
        - Temperatura: {temp_actual}°C (sin riesgo de hielo)
        - Sin lluvia (probabilidad: {prob_lluvia}%)
        - Sin nieve (probabilidad: {prob_nieve}%)
        - Viento normal: {viento} km/h
        - Buena visibilidad: {visibilidad} km
        
        ¿Cuáles son los procedimientos operativos normales para estas condiciones? 
        IMPORTANTE: NO se requieren procedimientos especiales para hielo, nieve, lluvia o viento fuerte ya que NO se presentan estas condiciones.
        """
    
    return contexto_climatico + consulta_especifica

def procesar_consulta_clima_procedimientos(fecha: str, ciudad: str = "rio grande") -> Dict[str, Any]:
    """Función principal que procesa la consulta completa"""
    resultado = {
        'fecha': fecha,
        'ciudad': ciudad,
        'clima_obtenido': False,
        'procedimientos_consultados': False,
        'respuesta_generada': False,
        'workflow_disparado': False,
        'datos_clima': None,
        'condiciones_analizadas': None,
        'respuesta_llm': None,
        'fuentes': [],
        'workflow_info': None
    }
    
    try:
        # 1. Obtener datos del clima
        datos_clima = obtener_clima(ciudad, fecha)
        
        if not datos_clima:
            return resultado
        
        resultado['clima_obtenido'] = True
        resultado['datos_clima'] = datos_clima
        
        # 2. Analizar condiciones climáticas
        condiciones = analizar_condiciones_climaticas(datos_clima)
        resultado['condiciones_analizadas'] = condiciones
        
        # 3. Verificar si hay condiciones adversas
        condiciones_adversas = condiciones.get('condiciones_adversas', [])
        
        if not condiciones_adversas:
            # Condiciones normales - NO se requieren procedimientos ni tareas
            resultado['procedimientos_consultados'] = True
            resultado['respuesta_llm'] = "Las condiciones meteorológicas son normales. No se requieren procedimientos especiales ni tareas de control de hielo/nieve."
            resultado['fuentes'] = []
            resultado['respuesta_generada'] = True
        else:
            # 4. Generar consulta para RAG (solo si hay condiciones adversas)
            consulta_rag = generar_consulta_rag(condiciones)
            
            # 5. Consultar sistema RAG
            rag_service = RAGService()
            respuesta, fuentes = rag_service.answer_question(consulta_rag)
            
            resultado['procedimientos_consultados'] = True
            resultado['respuesta_llm'] = respuesta
            resultado['fuentes'] = fuentes
            resultado['respuesta_generada'] = True
        
        # 5. Enviar al workflow - DESACTIVADO TEMPORALMENTE
        # if respuesta:
        #     try:
        #         resultado_workflow = trigger_workflow("process-automation-service-binding.json", respuesta)
        #         
        #         if resultado_workflow:
        #             resultado['workflow_disparado'] = True
        #             resultado['workflow_info'] = resultado_workflow
        #         
        #     except Exception as e:
        #         logger.error(f"Error enviando al workflow: {e}")
        
        # 6. Guardar resultado en archivo JSON
        try:
            nombre_archivo = f"resultado_test_{fecha.replace('-', '')}.json"
            with open(nombre_archivo, 'w', encoding='utf-8') as f:
                json.dump(resultado, f, indent=2, ensure_ascii=False, default=str)
        except Exception as e:
            logger.warning(f"No se pudo guardar el archivo de resultado: {e}")
        
        return resultado
        
    except Exception as e:
        logger.error(f"Error en procesamiento completo: {e}")
        return resultado

# ==================== ENDPOINTS ====================

@app.get("/")
async def root():
    """Endpoint raíz - Información de la API"""
    return {
        "message": "API de Aeropuertos - Sistema de Análisis Climático y Procedimientos",
        "version": "1.0.0",
        "endpoints": {
            "docs": "/docs",
            "redoc": "/redoc",
            "health": "/health",
            "upload": "/upload",
            "ask": "/ask",
            "weather": "/weather",
            "stats": "/stats",
            "clear_all": "/clear_all"
        }
    }

@app.get("/health")
async def health_check():
    """Verificar el estado de la API"""
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat()
    }

@app.post("/upload", response_model=UploadResponse)
async def upload_document(file: UploadFile = File(...)):
    """Subir y procesar un documento"""
    try:
        # Verificar que el archivo no esté vacío
        if not file.filename:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No se seleccionó ningún archivo"
            )
        
        # Verificar extensión permitida
        if not allowed_file(file.filename):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Tipo de archivo no permitido. Extensiones permitidas: {', '.join(ALLOWED_EXTENSIONS)}"
            )
        
        # Leer contenido del archivo
        file_content = await file.read()
        
        # Guardar archivo en disco para descarga/listado
        try:
            dest_path = os.path.join(UPLOAD_FOLDER, file.filename)
            with open(dest_path, 'wb') as f:
                f.write(file_content)
        except Exception as e:
            logger.warning(f"No se pudo guardar el archivo en uploads: {e}")
        
        # Verificar tamaño
        if len(file_content) > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                detail=f"El archivo es demasiado grande. Tamaño máximo: {MAX_FILE_SIZE / 1024 / 1024}MB"
            )
        
        # Procesar documento
        rag_service = RAGService()
        success, message = rag_service.process_document(file.filename, file_content)
        
        return UploadResponse(success=success, message=message)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error procesando archivo: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error procesando archivo: {str(e)}"
        )

# ==================== DOCUMENTOS (LIST/DOWNLOAD) ====================

@app.get("/documents/list")
async def list_documents():
    """Listar documentos disponibles en la carpeta de uploads"""
    try:
        files = []
        for fname in os.listdir(UPLOAD_FOLDER):
            fpath = os.path.join(UPLOAD_FOLDER, fname)
            if os.path.isfile(fpath):
                size_kb = round(os.path.getsize(fpath) / 1024.0, 2)
                files.append({
                    "filename": fname,
                    "size_kb": size_kb
                })
        return {"success": True, "documents": files}
    except Exception as e:
        logger.error(f"Error listando documentos: {e}")
        return {"success": False, "message": f"Error: {str(e)}"}

@app.get("/documents/download/{filename}")
async def download_document(filename: str):
    """Descargar un documento almacenado en uploads"""
    try:
        safe_name = os.path.basename(filename)
        fpath = os.path.join(UPLOAD_FOLDER, safe_name)
        if not os.path.isfile(fpath):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Archivo no encontrado")
        return FileResponse(fpath, media_type="application/octet-stream", filename=safe_name)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error descargando documento: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Error: {str(e)}")

@app.post("/ask", response_model=QuestionResponse)
async def ask_question(request: QuestionRequest):
    """Hacer una pregunta al chatbot RAG"""
    try:
        rag_service = RAGService()
        answer, sources = rag_service.answer_question(request.question)
        
        return QuestionResponse(
            success=True,
            answer=answer,
            sources=sources
        )
        
    except Exception as e:
        logger.error(f"Error en pregunta: {e}")
        return QuestionResponse(
            success=False,
            message=f"Error: {str(e)}"
        )

@app.post("/weather", response_model=WeatherResponse)
async def get_weather(request: WeatherRequest):
    """Obtener análisis climático y procedimientos"""
    try:
        # Validar fecha
        fecha_obj = datetime.strptime(request.fecha, '%Y-%m-%d').date()
        fecha_actual = date.today()
        diferencia_dias = (fecha_obj - fecha_actual).days
        
        if diferencia_dias < -30:
            return WeatherResponse(
                success=False,
                message='La fecha es muy antigua. Ingrese una fecha más reciente.'
            )
        elif diferencia_dias > 14:
            return WeatherResponse(
                success=False,
                message='La fecha es muy lejana en el futuro. Ingrese una fecha dentro de los próximos 14 días.'
            )
        
        # Procesar consulta completa
        resultado_completo = procesar_consulta_clima_procedimientos(request.fecha, request.ciudad)
        
        if resultado_completo and resultado_completo['clima_obtenido']:
            return WeatherResponse(
                success=True,
                resultado=resultado_completo
            )
        else:
            return WeatherResponse(
                success=False,
                message='Error obteniendo datos del clima'
            )
            
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Formato de fecha inválido. Use YYYY-MM-DD"
        )
    except Exception as e:
        logger.error(f"Error procesando clima: {e}")
        return WeatherResponse(
            success=False,
            message=f"Error: {str(e)}"
        )

@app.get("/stats", response_model=StatsResponse)
async def get_stats():
    """Obtener estadísticas de la base de datos"""
    try:
        rag_service = RAGService()
        stats = rag_service.get_stats()
        
        if stats:
            return StatsResponse(success=True, stats=stats)
        else:
            return StatsResponse(
                success=False,
                message='No se pudieron obtener estadísticas'
            )
            
    except Exception as e:
        logger.error(f"Error obteniendo estadísticas: {e}")
        return StatsResponse(
            success=False,
            message=f"Error: {str(e)}"
        )

@app.delete("/clear_all", response_model=ClearResponse)
async def clear_all():
    """Eliminar todos los documentos de la base de datos"""
    try:
        rag_service = RAGService()
        success, message = rag_service.clear_all_documents()
        
        return ClearResponse(success=success, message=message)
        
    except Exception as e:
        logger.error(f"Error limpiando documentos: {e}")
        return ClearResponse(
            success=False,
            message=f"Error: {str(e)}"
        )

@app.post("/historico", response_model=HistoricoResponse)
async def consultar_historico(request: HistoricoRequest):
    """Consultar datos de tareas por rango de fechas desde tareas.xlsx"""
    try:
        # Validar fechas
        try:
            fecha_inicio_obj = datetime.strptime(request.fecha_inicio, '%Y-%m-%d').date()
            fecha_fin_obj = datetime.strptime(request.fecha_fin, '%Y-%m-%d').date()
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Formato de fecha inválido. Use YYYY-MM-DD"
            )
        
        if fecha_inicio_obj > fecha_fin_obj:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="La fecha de inicio debe ser anterior a la fecha de fin"
            )
        
        # Leer archivo Excel
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Librería openpyxl no instalada"
            )
        
        excel_path = os.path.join('data', 'tareas.xlsx')
        if not os.path.exists(excel_path):
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Archivo de tareas no encontrado"
            )
        
        datos_filtrados = []
        
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            ws = wb.active
            
            # Obtener headers de la fila 2 (índice 2)
            headers = [cell.value for cell in ws[2]]
            
            # Leer datos desde la fila 3
            for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
                try:
                    # La fecha está en la columna A (índice 0)
                    fecha_cell = row[0]
                    
                    if fecha_cell is None:
                        continue
                    
                    # Convertir a date si es datetime
                    if isinstance(fecha_cell, datetime):
                        fecha_registro = fecha_cell.date()
                    elif isinstance(fecha_cell, date):
                        fecha_registro = fecha_cell
                    else:
                        # Intentar parsear como string
                        fecha_registro = datetime.strptime(str(fecha_cell), '%Y-%m-%d').date()
                    
                    # Filtrar por rango de fechas
                    if fecha_inicio_obj <= fecha_registro <= fecha_fin_obj:
                        registro = {
                            'id': row_idx,
                            'fecha': fecha_registro.strftime('%Y-%m-%d'),
                            'salida_sol': str(row[1]) if row[1] is not None else '',
                            'puesta_sol': str(row[2]) if row[2] is not None else '',
                            'motivo_activacion': str(row[3]) if row[3] is not None else '',
                            'fenomeno_meteorologico': str(row[4]) if row[4] is not None else '',
                            'tipo_trabajo': str(row[5]) if row[5] is not None else '',
                            'vehiculo': str(row[6]) if row[6] is not None else '',
                            'equipo': str(row[7]) if row[7] is not None else '',
                            'urea_kilos': row[8] if row[8] is not None else '',
                            'glicol_litros': row[9] if row[9] is not None else '',
                            'prioridad_trabajo_1': str(row[10]) if row[10] is not None else '',
                            'prioridad_trabajo_2': str(row[11]) if row[11] is not None else '',
                            'temperatura': row[19] if row[19] is not None else '',
                            'punto_rocio': row[20] if row[20] is not None else '',
                            'humedad': row[21] if row[21] is not None else '',
                            'presion': row[22] if row[22] is not None else '',
                            'viento': row[23] if row[23] is not None else ''
                        }
                        datos_filtrados.append(registro)
                        
                        # Limitar resultados
                        if len(datos_filtrados) >= request.limite:
                            break
                            
                except (ValueError, IndexError) as e:
                    # Saltar registros con errores
                    logger.warning(f"Error procesando fila {row_idx}: {e}")
                    continue
            
            wb.close()
            
        except Exception as e:
            logger.error(f"Error leyendo archivo Excel: {e}")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Error leyendo archivo Excel: {str(e)}"
            )
        
        return HistoricoResponse(
            success=True,
            datos=datos_filtrados,
            total_registros=len(datos_filtrados)
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error consultando tareas: {e}")
        return HistoricoResponse(
            success=False,
            message=f"Error: {str(e)}"
        )

# ==================== AVISOS ====================

@app.post("/generar-avisos")
async def generar_avisos_endpoint(condiciones: Dict[str, Any]):
    """Genera avisos basados en las condiciones climáticas"""
    try:
        resultado_avisos = generar_avisos(condiciones)
        
        # Agregar tareas del procedimiento para cada aviso
        for aviso in resultado_avisos.get('avisos_generados', []):
            tipo_aviso = aviso.get('tipo', '')
            aviso['tareas_procedimiento'] = obtener_tareas_procedimiento(tipo_aviso)
        
        return {
            "success": True,
            "avisos": resultado_avisos
        }
    except Exception as e:
        logger.error(f"Error generando avisos: {e}")
        return {
            "success": False,
            "message": f"Error: {str(e)}"
        }

# ==================== SIMULACIÓN ====================

@app.get("/simulacion/escenarios")
async def listar_escenarios():
    """Lista los escenarios de simulación disponibles"""
    try:
        return obtener_escenarios_disponibles()
    except Exception as e:
        logger.error(f"Error listando escenarios: {e}")
        return {
            "success": False,
            "message": f"Error: {str(e)}"
        }

@app.get("/simulacion/{escenario}")
async def obtener_simulacion(escenario: str):
    """Obtiene datos simulados para el escenario especificado"""
    try:
        resultado = generar_datos_simulados(escenario)
        return {"resultado": resultado}
    except Exception as e:
        logger.error(f"Error generando simulación: {e}")
        return {
            "success": False,
            "message": f"Error: {str(e)}"
        }

# ==================== SAP INTEGRATION SUITE ====================

class EnviarAvisoISuiteRequest(BaseModel):
    """Request para enviar aviso a SAP Integration Suite"""
    aviso: Dict[str, Any] = Field(..., description="JSON del aviso ya generado por el sistema")

class EnviarAvisoISuiteResponse(BaseModel):
    """Response del envío a SAP Integration Suite"""
    success: bool
    status_code: Optional[int] = None
    response_body: Optional[Any] = None
    message: str
    timestamp: str

class ISuiteStatusResponse(BaseModel):
    """Response del estado de configuración de SAP Integration Suite"""
    success: bool
    configuracion: Optional[Dict[str, Any]] = None
    token_status: Optional[Dict[str, Any]] = None
    message: str

@app.post("/enviar-aviso-isuite", response_model=EnviarAvisoISuiteResponse)
async def enviar_aviso_isuite_endpoint(request: EnviarAvisoISuiteRequest):
    """
    Envía un aviso ya generado a SAP Integration Suite.
    
    Este endpoint recibe el aviso completo generado por el sistema y lo envía
    al iFlow HTTP de SAP Integration Suite usando OAuth2 Client Credentials.
    
    El aviso debe tener la estructura retornada por /generar-avisos.
    """
    try:
        logger.info("Recibida solicitud para enviar aviso a SAP Integration Suite")
        
        # Validar que el aviso no esté vacío
        if not request.aviso:
            return EnviarAvisoISuiteResponse(
                success=False,
                message="El aviso está vacío",
                timestamp=datetime.now().isoformat()
            )
        
        # Llamar al módulo isuite.py
        resultado = enviar_aviso_a_isuite(request.aviso)
        
        return EnviarAvisoISuiteResponse(
            success=resultado.get('success', False),
            status_code=resultado.get('status_code'),
            response_body=resultado.get('response_body'),
            message=resultado.get('message', ''),
            timestamp=resultado.get('timestamp', datetime.now().isoformat())
        )
        
    except Exception as e:
        logger.error(f"Error enviando aviso a SAP Integration Suite: {e}")
        return EnviarAvisoISuiteResponse(
            success=False,
            message=f"Error interno: {str(e)}",
            timestamp=datetime.now().isoformat()
        )

@app.get("/isuite/status", response_model=ISuiteStatusResponse)
async def isuite_status():
    """
    Verifica el estado de la configuración de SAP Integration Suite.
    
    Útil para diagnóstico y verificar que las variables de entorno
    estén correctamente configuradas antes de intentar enviar avisos.
    """
    try:
        # Verificar configuración
        config_status = verificar_config_isuite()
        
        # Obtener estado del token
        token_status = obtener_estado_token()
        
        return ISuiteStatusResponse(
            success=config_status.get('configuracion_valida', False),
            configuracion=config_status,
            token_status=token_status,
            message=config_status.get('mensaje', '')
        )
        
    except Exception as e:
        logger.error(f"Error verificando estado de SAP Integration Suite: {e}")
        return ISuiteStatusResponse(
            success=False,
            message=f"Error: {str(e)}"
        )

# ==================== STATION DATA (MARWIS) ====================

@app.get("/station-data", response_model=StationSensorsResponse)
async def get_station_data():
    """Retorna la lista de sensores desde backend/station_data.json"""
    try:
        json_path = os.path.join(os.path.dirname(__file__), "station_data.json")
        if not os.path.exists(json_path):
            return StationSensorsResponse(success=False, message="station_data.json no encontrado")
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            return StationSensorsResponse(success=True, sensors=data, total=len(data))
        else:
            # Compatibilidad con formato antiguo
            sensors = data.get("measurements") or []
            return StationSensorsResponse(success=True, sensors=sensors, total=len(sensors))
    except Exception as e:
        logger.error(f"Error leyendo station_data.json: {e}")
        return StationSensorsResponse(success=False, message=f"Error: {str(e)}")

@app.post("/station-data/refresh", response_model=StationSensorsResponse)
async def refresh_station_data():
    """Consulta la API de ViewMondo y devuelve la lista actualizada de sensores"""
    try:
        data = run_marwis()
        return StationSensorsResponse(success=True, sensors=data, total=len(data))
    except Exception as e:
        logger.error(f"Error ejecutando MARWIS: {e}")
        return StationSensorsResponse(success=False, message=f"Error: {str(e)}")

# ==================== EJECUCIÓN ====================

if __name__ == '__main__':
    import uvicorn
    uvicorn.run(
        "fastapi_app:app",
        host="0.0.0.0",
        port=8000,
        reload=True,
        log_level="info"
    )
